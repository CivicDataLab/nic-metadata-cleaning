"""Build Dublin Core mapping sheet from DuckDB metadata tables."""
import argparse
import datetime as _dt
import duckdb
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import csv

DEFAULT_RAW_XLSX = "/home/aakash/NIC/Newfolder/ResourceList_DoFHW/ResourceList_Department-of-Health-and-Family-Welfare.xlsx"

# (display label, raw_metadata column expression or None, dublin_core_metadata column or None)
SCHEMA = {
    "Exact Match": [
        ("Title",                                "Title",                                    "Title"),
        ("Subject[sector_resource]",             "sector_resource",                          "Subject[sector_resource]"),
        ("Date Created",                         "created",                                  "Date Created"),
        ("Date Issued",                          "published",                                "Date Issued"),
        ("Date Modified",                        "changed",                                  "Date Modified"),
        ("Type",                                 "resource_category",                        "Type"),
        ("Extent",                               "file_size",                                "Extent"),
        ("Format (or MediaType)",                "file_format",                              "Format"),
        ("Relation",                             '"Reference Url"',                          "Relation"),
        ("Identifier[UUID]",                     "uuid",                                     "Identifier[UUID]"),
        ("Identifier[landing page]",             "domain || '/' || node_alias",              "Identifier[landing_page]"),
        ("Source",                               None,                                       "Source"),
        ("Spatial Coverage",                     None,                                       "Spatial Coverage"),
        ("Temporal Coverage",                    None,                                       "Temporal Coverage"),
        ("Jurisdiction",                         "govt_type",                                "Jurisdiction"),
        ("Rights Statement",                     None,                                       "Rights Statement"),
        ("Access Rights",                        '"Access Type"',                            "Access Rights"),
        ("Accrual Periodicity",                  "frequency",                                "Accrual Periodicity"),
        ("Coverage",                             "granularity",                              "Coverage"),
        ("Subject [keyword]",                    None,                                       "Subject[keyword]"),
        ("Subject [sponsored_keyword]",          None,                                       "Subject[sponsored_keyword]"),
        ("Subject[sector]",                      "sector",                                   "Subject[sector]"),
        ("Relation [download_url]",              "datafile",                                 "Relation[download_url]"),
        ("Relation [Access URL]",                None,                                       "Relation[Access URL]"),
        ("Relation [endpointURL]",               "datafile_url",                             "Relation[endpointURL]"),
        ("Depiction (foaf:depiction)",           None,                                       "Depiction"),
        ("Relation[Catalog Title]",              "catalog_title",                            "Relation[Catalog Title]"),
        ("Conforms To",                          None,                                       "Conforms To"),
    ],
    "AMBIGIOUS": [
        ("Rights",                               None,                                       "Rights"),
        ("High Value Dataset Category",          "field_high_value_dataset",                 "High Value Dataset Category"),
        ("Accrual Method",                       "field_resource_type",                      "Accrual Method"),
        ("Creator",                              "cdos_state_ministry",                      "Creator"),
        ("Publisher[state_department]",          "state_department",                         "Publisher[state_department]"),
        ("Publisher[ministry_department]",       "ministry_department",                      "Publisher[ministry_department]"),
        ("Note",                                 "note",                                     "Note"),
    ],
    "GAPS": [
        ("Description",                          None,                                       "Description"),
        ("License",                              None,                                       "License"),
        ("Language",                             None,                                       "Language"),
        ("Abstract",                             None,                                       "Abstract"),
        ("Alternative Title",                    None,                                       "Alternative Title"),
        ("Relation [Has Part]",                  None,                                       "Relation[Has Part]"),
        ("Relation [Is Referenced By]",          None,                                       "Relation[Is Referenced By]"),
        ("Relation [Is Replaced By]",            None,                                       "Relation[Is Replaced By]"),
        ("Relation[Replaces]",                   None,                                       "Relation[Replaces]"),
        ("Relation[Is Version Of]",              None,                                       "Relation[Is Version Of]"),
        ("Relation[Has Version]",                None,                                       "Relation[Has Version]"),
        ("Description[endpointDescription]",     None,                                       "Description[endpointDescription]"),
        ("Collection",                           None,                                       "Collection"),
        ("Relation[Is Part Of]",                 None,                                       "Relation[Is Part Of]"),
    ],
    "Unmappable": [
        ("ogdp_view_count",                      "ogdp_view_count",                          "ogdp_view_count"),
        ("ogdp_download_count",                  "ogdp_download_count",                      "ogdp_download_count"),
        ("api_request_count",                    "api_request_count",                        "api_request_count"),
        ("field_show_export",                    "field_show_export",                        "field_show_export"),
        ("is_rated",                             "is_rated",                                 "is_rated"),
        ("external_api_reference",               "external_api_reference",                   "external_api_reference"),
        ("is_api_available",                     "is_api_available",                         "is_api_available"),
        ("is_visualized",                        "is_visualized",                            "is_visualized"),
        ("field_from_api",                       "field_from_api",                           "field_from_api"),
    ],
}

FLAGS = [
    ("PII Detected",                            "pii_detected"),
    ("PII Tested",                              "pii_tested"),
    ("PII test timestamp",                      "pii_test_timestamp"),
    ("Headers cleaned",                         "headers_cleaned"),
    ("Metadata enhanced",                       "metadata_enhanced"),
    ("Dataset merge",                           "dataset_merge"),
    ("Row count",                               "row_count"),
]


def normalize(v):
    if v is None:
        return ""
    try:
        import pandas as pd
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    return "" if s.lower() in ("nan", "nat", "none", "<na>") else s


def changed_flag(raw_val, dc_val):
    return 0 if normalize(raw_val) == normalize(dc_val) else 1


def _format_excel_value(v):
    """Render a value the way Excel displays it (dd/mm/yyyy for dates)."""
    if isinstance(v, _dt.datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
            return v.strftime("%d/%m/%Y")
        return v.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(v, _dt.date):
        return v.strftime("%d/%m/%Y")
    return v


def load_excel_raw(xlsx_path):
    """Return {uuid: {column_name: excel_formatted_value}} from the resource list."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    try:
        uuid_idx = headers.index("uuid")
    except ValueError:
        wb.close()
        raise RuntimeError(f"'uuid' column not found in {xlsx_path}")

    out = {}
    for row in rows_iter:
        uuid_val = row[uuid_idx]
        if uuid_val is None:
            continue
        record = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            record[h] = _format_excel_value(row[i])
        out[str(uuid_val).strip()] = record
    wb.close()
    return out


def evaluate_raw_expr(raw_expr, excel_row):
    """Resolve a SCHEMA raw_expr against an Excel row dict.

    Handles bare column names, quoted column names ("Reference Url"), and the
    `domain || '/' || node_alias` concatenation expression.
    """
    if not raw_expr:
        return ""
    if "||" in raw_expr:
        parts = []
        for piece in raw_expr.split("||"):
            piece = piece.strip()
            if piece.startswith("'") and piece.endswith("'"):
                parts.append(piece[1:-1])
            elif piece.startswith('"') and piece.endswith('"'):
                parts.append(str(excel_row.get(piece[1:-1], "") or ""))
            else:
                parts.append(str(excel_row.get(piece, "") or ""))
        return "".join(parts)
    if raw_expr.startswith('"') and raw_expr.endswith('"'):
        return excel_row.get(raw_expr[1:-1], "")
    return excel_row.get(raw_expr, "")


def build_query():
    """Build a SELECT against dublin_core_metadata only; raw values come from Excel."""
    select_parts = ['d."Identifier[UUID]" AS raw_uuid', 'd.batch AS batch']
    dc_aliases = {}

    for fields in SCHEMA.values():
        for _, _, dc_col in fields:
            if dc_col and dc_col not in dc_aliases:
                alias = f"dc_{len(dc_aliases)}"
                dc_aliases[dc_col] = alias
                select_parts.append(f'd."{dc_col}" AS {alias}')

    for _, dc_col in FLAGS:
        if dc_col not in dc_aliases:
            alias = f"dc_{len(dc_aliases)}"
            dc_aliases[dc_col] = alias
            select_parts.append(f'd."{dc_col}" AS {alias}')

    sql = f"""
        SELECT {', '.join(select_parts)}
        FROM dublin_core_metadata d
        WHERE d.batch BETWEEN ? AND ?
        ORDER BY d.batch, d."Identifier[UUID]"
    """
    return sql, dc_aliases


def fetch_datasets(db_path, from_batch, to_batch):
    sql, dc_aliases = build_query()
    con = duckdb.connect(db_path, read_only=True)
    df = con.execute(sql, [from_batch, to_batch]).fetchdf()
    con.close()
    return df, dc_aliases


def build_rows(df, excel_data, dc_aliases):
    n = len(df)
    header1 = ["Dublin Core Element", "Equivalent OGD Dataset Metadata"]
    header2 = ["", ""]
    for i in range(n):
        label = f"Dataset {i + 1} ({df.iloc[i]['raw_uuid']})"
        header1 += [label, label, label]
        header2 += ["Current Mapped OGD Metadata", "New Corrected Metadata Value", "Changed (0 for no 1 for yes)"]

    out = [header1, header2]

    excel_rows = [excel_data.get(str(df.iloc[i]["raw_uuid"]).strip(), {}) for i in range(n)]

    for section, fields in SCHEMA.items():
        out.append([section] + [""] * (len(header1) - 1))
        for label, raw_expr, dc_col in fields:
            ogd_display = raw_expr.replace('"', '') if raw_expr else ""
            row = [label, ogd_display]
            dc_alias = dc_aliases.get(dc_col) if dc_col else None
            for i in range(n):
                raw_val = evaluate_raw_expr(raw_expr, excel_rows[i]) if raw_expr else ""
                dc_val = df.iloc[i][dc_alias] if dc_alias else ""
                row += [normalize(raw_val), normalize(dc_val), changed_flag(raw_val, dc_val)]
            out.append(row)

    out.append([""] * len(header1))
    out.append(["", "Flags"] + [""] * (len(header1) - 2))
    for flag_label, flag_col in FLAGS:
        alias = dc_aliases.get(flag_col)
        row = ["", flag_label]
        for i in range(n):
            val = normalize(df.iloc[i][alias]) if alias else ""
            row += [val, "", ""]
        out.append(row)

    return out


def write_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


XLSX_CELL_LIMIT = 32767  # Excel max characters per cell


def _truncate_for_xlsx(v):
    """Excel limits cells to 32767 chars; truncate cleanly at a comma boundary."""
    if not isinstance(v, str) or len(v) <= XLSX_CELL_LIMIT:
        return v
    marker = " …[truncated]"
    cut = XLSX_CELL_LIMIT - len(marker)
    last_comma = v.rfind(",", 0, cut)
    if last_comma > 0:
        cut = last_comma
    return v[:cut] + marker


def write_xlsx(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping"
    for row in rows:
        ws.append([_truncate_for_xlsx(c) for c in row])

    bold = Font(bold=True, name="Arial")
    section_fill = PatternFill("solid", start_color="FFE699")
    header_fill = PatternFill("solid", start_color="BDD7EE")
    section_names = set(SCHEMA.keys())

    for r_idx, row in enumerate(rows, start=1):
        first = row[0] if row else ""
        second = row[1] if len(row) > 1 else ""
        is_header = r_idx <= 2
        is_section = first in section_names or second == "Flags"
        for cell in ws[r_idx]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if is_header:
                cell.font = bold
                cell.fill = header_fill
            elif is_section:
                cell.font = bold
                cell.fill = section_fill
            else:
                cell.font = Font(name="Arial")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32
    for col in range(3, len(rows[0]) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 40
    ws.freeze_panes = "C3"
    wb.save(path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--db", default="transformation/metadata.db")
    p.add_argument("--raw-xlsx", default=DEFAULT_RAW_XLSX,
                   help="Resource list xlsx used for 'before' values")
    p.add_argument("--from-batch", type=int, required=True)
    p.add_argument("--to-batch", type=int, required=True)
    p.add_argument("--format", choices=["xlsx", "csv"], default="xlsx")
    p.add_argument("--out", default=None)
    args = p.parse_args()

    df, dc_aliases = fetch_datasets(args.db, args.from_batch, args.to_batch)
    if len(df) == 0:
        print(f"No datasets found in batches {args.from_batch}-{args.to_batch}")
        return

    print(f"Loading raw values from {args.raw_xlsx} ...")
    excel_data = load_excel_raw(args.raw_xlsx)
    print(f"Loaded {len(excel_data)} resource rows from xlsx")

    rows = build_rows(df, excel_data, dc_aliases)
    out_path = args.out or f"Dublin_Mapping_batches_{args.from_batch}_to_{args.to_batch}.{args.format}"

    if args.format == "csv":
        write_csv(rows, out_path)
    else:
        write_xlsx(rows, out_path)

    print(f"Wrote {len(df)} datasets to {out_path}")


if __name__ == "__main__":
    main()
    