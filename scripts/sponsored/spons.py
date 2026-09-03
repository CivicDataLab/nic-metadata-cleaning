# import re
# import math
# import argparse
# from collections import Counter
# from pathlib import Path

# import openpyxl
# from tqdm import tqdm


# # ============================================================
# # CONFIGURATION
# # ============================================================

# DEFAULT_INPUT = "input1.xlsx"
# DEFAULT_OUTPUT = "sponsored_keywords_output2.xlsx"

# KEYWORD_COLUMN = "generated_keywords"
# SPONSORED_COLUMN = "sponsored keywords"

# # Maximum sponsored keywords per dataset
# MAX_SPONSORED = 2


# # ============================================================
# # HIGH-PRIORITY SHORT FORMS / ACRONYMS
# # ============================================================
# #
# # These are given the highest priority when they occur in
# # generated_keywords.
# #
# # You can keep adding acronyms to this set as your controlled
# # vocabulary grows.
# #

# PRIORITY_ACRONYMS = {
#     # Health surveys / systems
#     "NFHS",
#     "DLHS",
#     "AHS",
#     "HMIS",
#     "RCH",
#     "SRS",
#     "CRS",
#     "MCTS",
#     "ANM",
#     "ASHA",
#     "CHC",
#     "PHC",
#     "UPHC",
#     "HWC",

#     # Health / family welfare
#     "ICDS",
#     "NRHM",
#     "NHM",
#     "NHP",
#     "NSS",
#     "NSSO",
#     "IIPS",
#     "WHO",
#     "UNICEF",
#     "UNFPA",
#     "UNDP",

#     # Census / demographic sources
#     "CENSUS",
#     "ORGI",
#     "CRS",

#     # Government / administrative
#     "UIDAI",
#     "Aadhaar",
#     "NITI",
#     "NITI AAYOG",

#     # Schemes / programmes
#     "PMJAY",
#     "PMAY",
#     "MGNREGA",
#     "NREGA",
#     "JSY",
#     "JSSK",
#     "RKSK",
#     "RMNCH",
#     "RMNCHA",

#     # Other common government datasets
#     "GDP",
#     "CPI",
#     "WPI",
#     "IIP",
#     "MSME",
#     "GST",
#     "TDS",
#     "PAN",
#     "EPFO",
#     "ESIC",
#     "NABARD",
#     "SEBI",
#     "RBI",
#     "NPCI",

#     # Education
#     "UDISE",
#     "UDISE+",
#     "DISE",
#     "CBSE",
#     "NCERT",
#     "AISHE",
#     "NSS",
#     "GER",
#     "NER",

#     # Agriculture
#     "PMFBY",
#     "KCC",
#     "FPO",
#     "APMC",

#     # Urban / local government
#     "ULB",
#     "NULM",
#     "AMRUT",
#     "MLCP",
#     "SWM",

#     # Transport
#     "IPT",
#     "MRTS",
#     "NHAI",
#     "RTO",

#     # Environment
#     "CPCB",
#     "SPCB",
#     "NGT",

#     # Other
#     "GIS",
#     "GPS",
#     "API",
#     "AI",
#     "ML",
#     "NLP",
# }


# # Normalized version for fast lookup
# PRIORITY_ACRONYMS_NORMALIZED = {
#     x.lower()
#     for x in PRIORITY_ACRONYMS
# }


# # ============================================================
# # BLOCKED SPONSORED KEYWORDS
# # ============================================================
# #
# # These words should NEVER be selected as sponsored keywords.
# #
# # This specifically includes the words you mentioned:
# # population, demography
# #
# # Add more generic terms here whenever required.
# #

# BLOCKED_SPONSORED_KEYWORDS = {
#     # Demography / population
#     "population",
#     "populations",
#     "demography",
#     "demographic",
#     "demographics",

#     # Generic health terms
#     "health",
#     "healthcare",
#     "medical",
#     "medicine",
#     "disease",
#     "diseases",

#     # Generic gender / population terms
#     "women",
#     "woman",
#     "men",
#     "man",
#     "children",
#     "child",
#     "people",
#     "persons",
#     "person",
#     "household",
#     "households",

#     # Generic dataset terms
#     "data",
#     "dataset",
#     "datasets",
#     "information",
#     "records",
#     "record",
#     "details",

#     # Generic publication terms
#     "report",
#     "reports",
#     "survey",
#     "surveys",
#     "study",
#     "studies",
#     "analysis",
#     "statistical",
#     "statistics",
#     "statistic",

#     # Time-related generic terms
#     "annual",
#     "year",
#     "yearly",
#     "monthly",
#     "quarterly",
#     "weekly",
#     "daily",

#     # Generic analytical terms
#     "indicator",
#     "indicators",
#     "comparison",
#     "comparisons",
#     "status",
#     "value",
#     "values",
#     "number",
#     "numbers",
#     "total",
#     "distribution",
#     "percentage",

#     # Metadata-type terms
#     "description",
#     "descriptions",
#     "variable",
#     "variables",
#     "parameter",
#     "parameters",
#     "schedule",
#     "section",
#     "table",
#     "tables",

#     # Generic geography
#     "india",
#     "state",
#     "states",
#     "district",
#     "districts",
#     "national",
#     "central",
#     "regional",
#     "region",

#     # Generic government terms
#     "government",
#     "ministry",
#     "department",
#     "scheme",
#     "schemes",
#     "programme",
#     "program",
#     "service",
#     "services",

#     # Generic words
#     "facility",
#     "facilities",
#     "coverage",
#     "information",
# }


# # ============================================================
# # NORMALIZATION
# # ============================================================

# def normalize_keyword(keyword):
#     """
#     Normalize keyword for comparison.

#     Example:
#         " NFHS " -> "nfhs"
#         "UNICEF CES" -> "unicef ces"
#     """

#     if keyword is None:
#         return ""

#     keyword = str(keyword).strip()

#     keyword = re.sub(
#         r"\s+",
#         " ",
#         keyword
#     )

#     return keyword.lower()


# # ============================================================
# # SPLIT KEYWORDS
# # ============================================================

# def split_keywords(value):
#     """
#     Convert comma-separated generated_keywords into a list.

#     Example:
#         "AHS, annual, health, survey"
#         ->
#         ["AHS", "annual", "health", "survey"]
#     """

#     if value is None:
#         return []

#     value = str(value).strip()

#     if not value:
#         return []

#     parts = [
#         x.strip()
#         for x in value.split(",")
#     ]

#     return [
#         x
#         for x in parts
#         if x
#     ]


# # ============================================================
# # ACRONYM DETECTION
# # ============================================================

# def is_priority_acronym(keyword):
#     """
#     Check whether a keyword is in the controlled
#     priority acronym list.

#     This is stronger than simply checking whether
#     something is uppercase.
#     """

#     normalized = normalize_keyword(keyword)

#     return (
#         normalized
#         in PRIORITY_ACRONYMS_NORMALIZED
#     )


# def looks_like_acronym(keyword):
#     """
#     Detect additional acronym-like keywords.

#     Examples:
#         NFHS
#         DLHS
#         HMIS
#         RCH
#         SRS
#         UDISE+
#     """

#     clean = re.sub(
#         r"[^A-Za-z0-9+]",
#         "",
#         str(keyword)
#     )

#     if not clean:
#         return False

#     # Explicit controlled acronym
#     if is_priority_acronym(keyword):
#         return True

#     # Generic uppercase acronym detection
#     if (
#         len(clean) >= 2
#         and clean.upper() == clean
#         and any(c.isalpha() for c in clean)
#     ):
#         return True

#     return False


# # ============================================================
# # BLOCKED KEYWORD CHECK
# # ============================================================

# def is_blocked(keyword):
#     """
#     Return True if keyword should never be sponsored.
#     """

#     normalized = normalize_keyword(keyword)

#     if normalized in BLOCKED_SPONSORED_KEYWORDS:
#         return True

#     return False


# # ============================================================
# # GLOBAL KEYWORD FREQUENCY
# # ============================================================

# def calculate_keyword_frequency(
#     input_file,
#     keyword_col
# ):
#     """
#     First pass through the Excel file.

#     Counts how many datasets contain each keyword.
#     """

#     print()
#     print("=" * 70)
#     print("PASS 1/2 - CALCULATING KEYWORD FREQUENCY")
#     print("=" * 70)

#     wb = openpyxl.load_workbook(
#         input_file,
#         read_only=True,
#         data_only=True
#     )

#     ws = wb.active

#     header_row = next(
#         ws.iter_rows(
#             min_row=1,
#             max_row=1,
#             values_only=True
#         )
#     )

#     headers = list(header_row)

#     if keyword_col not in headers:
#         raise ValueError(
#             f"\nColumn '{keyword_col}' not found.\n"
#             f"Available columns:\n{headers}"
#         )

#     keyword_idx = headers.index(
#         keyword_col
#     )

#     frequency = Counter()

#     total_rows = 0

#     for row in tqdm(
#         ws.iter_rows(
#             min_row=2,
#             values_only=True
#         ),
#         desc="Scanning datasets"
#     ):

#         total_rows += 1

#         value = row[keyword_idx]

#         keywords = split_keywords(
#             value
#         )

#         # A keyword should count only once
#         # within the same dataset.
#         seen = set()

#         for keyword in keywords:

#             normalized = normalize_keyword(
#                 keyword
#             )

#             if (
#                 normalized
#                 and normalized not in seen
#             ):
#                 frequency[normalized] += 1
#                 seen.add(normalized)

#     wb.close()

#     print()
#     print(f"Total datasets : {total_rows:,}")
#     print(
#         f"Unique keywords: {len(frequency):,}"
#     )

#     return frequency, total_rows


# # ============================================================
# # SEMANTIC SCORE
# # ============================================================

# def semantic_score(
#     keyword,
#     frequency,
#     total_rows
# ):
#     """
#     Score keyword for sponsored-keyword suitability.

#     Priority order:

#     1. Controlled acronym
#     2. Other acronym
#     3. Rare / specific keyword
#     4. Longer / multi-word keyword

#     Generic words are blocked completely.
#     """

#     normalized = normalize_keyword(
#         keyword
#     )

#     if not normalized:
#         return -999999

#     # --------------------------------------------------------
#     # NEVER SPONSOR BLOCKED WORDS
#     # --------------------------------------------------------

#     if is_blocked(keyword):
#         return -999999

#     # --------------------------------------------------------
#     # HIGHEST PRIORITY:
#     # CONTROLLED ACRONYMS
#     # --------------------------------------------------------

#     if is_priority_acronym(keyword):

#         # Very high score ensures these are selected
#         # before normal semantic keywords.
#         return 10000

#     # --------------------------------------------------------
#     # OTHER ACRONYMS
#     # --------------------------------------------------------

#     if looks_like_acronym(keyword):

#         return 8000

#     # --------------------------------------------------------
#     # FREQUENCY / RARITY
#     # --------------------------------------------------------

#     freq = frequency.get(
#         normalized,
#         1
#     )

#     rarity = math.log(
#         (total_rows + 1)
#         /
#         (freq + 1)
#     )

#     # --------------------------------------------------------
#     # LENGTH / SPECIFICITY
#     # --------------------------------------------------------

#     character_bonus = min(
#         len(normalized) / 10,
#         3.0
#     )

#     word_count = len(
#         normalized.split()
#     )

#     multiword_bonus = min(
#         word_count * 0.7,
#         2.5
#     )

#     # --------------------------------------------------------
#     # SHORT WORD PENALTY
#     # --------------------------------------------------------

#     short_penalty = 0

#     if (
#         len(normalized) <= 3
#         and not looks_like_acronym(keyword)
#     ):
#         short_penalty = 3

#     # --------------------------------------------------------
#     # FINAL SCORE
#     # --------------------------------------------------------

#     score = (
#         rarity * 2.5
#         + character_bonus
#         + multiword_bonus
#         - short_penalty
#     )

#     return score


# # ============================================================
# # SELECT SPONSORED KEYWORDS
# # ============================================================

# def select_sponsored_keywords(
#     keywords,
#     frequency,
#     total_rows,
#     max_sponsored=2
# ):
#     """
#     Select up to 2 sponsored keywords.

#     Acronyms are always preferred.
#     Blocked generic words are excluded.
#     """

#     if not keywords:
#         return []

#     candidates = []

#     for position, keyword in enumerate(
#         keywords
#     ):

#         if is_blocked(keyword):
#             continue

#         score = semantic_score(
#             keyword,
#             frequency,
#             total_rows
#         )

#         if score <= -99999:
#             continue

#         candidates.append(
#             {
#                 "keyword": keyword,
#                 "score": score,
#                 "position": position,
#                 "is_priority_acronym":
#                     is_priority_acronym(
#                         keyword
#                     ),
#                 "is_acronym":
#                     looks_like_acronym(
#                         keyword
#                     ),
#             }
#         )

#     if not candidates:
#         return []

#     # --------------------------------------------------------
#     # SORTING
#     #
#     # Priority:
#     # 1. Controlled acronym
#     # 2. Other acronym
#     # 3. Semantic score
#     # 4. Original position
#     # --------------------------------------------------------

#     candidates.sort(
#         key=lambda x: (
#             x["is_priority_acronym"],
#             x["is_acronym"],
#             x["score"],
#             -x["position"]
#         ),
#         reverse=True
#     )

#     selected = []

#     selected_normalized = set()

#     for candidate in candidates:

#         keyword = candidate["keyword"]

#         normalized = normalize_keyword(
#             keyword
#         )

#         if normalized in selected_normalized:
#             continue

#         selected.append(keyword)

#         selected_normalized.add(
#             normalized
#         )

#         if len(selected) >= max_sponsored:
#             break

#     return selected


# # ============================================================
# # REMOVE SPONSORED KEYWORDS
# # ============================================================

# def remove_sponsored_keywords(
#     keywords,
#     sponsored
# ):
#     """
#     Remove sponsored keywords from
#     generated_keywords.
#     """

#     sponsored_normalized = {
#         normalize_keyword(x)
#         for x in sponsored
#     }

#     remaining = []

#     for keyword in keywords:

#         normalized = normalize_keyword(
#             keyword
#         )

#         if normalized not in sponsored_normalized:
#             remaining.append(keyword)

#     return remaining


# # ============================================================
# # CREATE OUTPUT
# # ============================================================

# def create_output(
#     input_file,
#     output_file,
#     keyword_col,
#     sponsored_col,
#     frequency,
#     total_rows
# ):
#     """
#     Second pass:
#     Create final Excel output.
#     """

#     print()
#     print("=" * 70)
#     print("PASS 2/2 - GENERATING SPONSORED KEYWORDS")
#     print("=" * 70)

#     input_wb = openpyxl.load_workbook(
#         input_file,
#         read_only=True,
#         data_only=True
#     )

#     input_ws = input_wb.active

#     header_row = next(
#         input_ws.iter_rows(
#             min_row=1,
#             max_row=1,
#             values_only=True
#         )
#     )

#     headers = list(header_row)

#     if keyword_col not in headers:
#         raise ValueError(
#             f"Column '{keyword_col}' not found."
#         )

#     keyword_idx = headers.index(
#         keyword_col
#     )

#     # --------------------------------------------------------
#     # OUTPUT HEADERS
#     # --------------------------------------------------------

#     output_headers = list(headers)

#     if sponsored_col not in output_headers:
#         output_headers.append(
#             sponsored_col
#         )

#     sponsored_idx = output_headers.index(
#         sponsored_col
#     )

#     # --------------------------------------------------------
#     # WRITE-ONLY WORKBOOK
#     # --------------------------------------------------------

#     output_wb = openpyxl.Workbook(
#         write_only=True
#     )

#     output_ws = output_wb.create_sheet(
#         "Output"
#     )

#     output_ws.append(
#         output_headers
#     )

#     # --------------------------------------------------------
#     # PROCESS DATASETS
#     # --------------------------------------------------------

#     sponsored_count = 0

#     for row in tqdm(
#         input_ws.iter_rows(
#             min_row=2,
#             values_only=True
#         ),
#         total=total_rows,
#         desc="Processing datasets"
#     ):

#         row = list(row)

#         generated_value = row[
#             keyword_idx
#         ]

#         keywords = split_keywords(
#             generated_value
#         )

#         # ----------------------------------------------------
#         # SELECT SPONSORED
#         # ----------------------------------------------------

#         sponsored = select_sponsored_keywords(
#             keywords=keywords,
#             frequency=frequency,
#             total_rows=total_rows,
#             max_sponsored=MAX_SPONSORED
#         )

#         # ----------------------------------------------------
#         # REMOVE SPONSORED FROM GENERATED KEYWORDS
#         # ----------------------------------------------------

#         remaining_keywords = (
#             remove_sponsored_keywords(
#                 keywords,
#                 sponsored
#             )
#         )

#         # ----------------------------------------------------
#         # UPDATE GENERATED KEYWORDS
#         # ----------------------------------------------------

#         row[keyword_idx] = ", ".join(
#             remaining_keywords
#         )

#         # ----------------------------------------------------
#         # UPDATE SPONSORED KEYWORDS
#         # ----------------------------------------------------

#         row[sponsored_idx] = ", ".join(
#             sponsored
#         )

#         if sponsored:
#             sponsored_count += 1

#         output_ws.append(row)

#     input_wb.close()

#     # --------------------------------------------------------
#     # SAVE
#     # --------------------------------------------------------

#     print()
#     print("Saving output file...")

#     output_wb.save(
#         output_file
#     )

#     print()
#     print("=" * 70)
#     print("COMPLETED")
#     print("=" * 70)

#     print(
#         f"Total datasets processed : "
#         f"{total_rows:,}"
#     )

#     print(
#         f"Datasets with sponsored  : "
#         f"{sponsored_count:,}"
#     )

#     print(
#         f"Output file              : "
#         f"{output_file}"
#     )


# # ============================================================
# # MAIN
# # ============================================================

# def main():

#     parser = argparse.ArgumentParser(
#         description=(
#             "Generate semantic sponsored keywords "
#             "from generated_keywords."
#         )
#     )

#     parser.add_argument(
#         "--input",
#         default=DEFAULT_INPUT,
#         help="Input Excel file"
#     )

#     parser.add_argument(
#         "--output",
#         default=DEFAULT_OUTPUT,
#         help="Output Excel file"
#     )

#     args = parser.parse_args()

#     input_file = Path(
#         args.input
#     )

#     output_file = Path(
#         args.output
#     )

#     if not input_file.exists():

#         raise FileNotFoundError(
#             f"Input file not found: "
#             f"{input_file}"
#         )

#     print()
#     print("=" * 70)
#     print("SPONSORED KEYWORD GENERATOR")
#     print("=" * 70)

#     print(
#         f"Input : {input_file}"
#     )

#     print(
#         f"Output: {output_file}"
#     )

#     print(
#         f"Maximum sponsored keywords: "
#         f"{MAX_SPONSORED}"
#     )

#     # --------------------------------------------------------
#     # PASS 1
#     # --------------------------------------------------------

#     frequency, total_rows = (
#         calculate_keyword_frequency(
#             input_file=input_file,
#             keyword_col=KEYWORD_COLUMN
#         )
#     )

#     # --------------------------------------------------------
#     # PASS 2
#     # --------------------------------------------------------

#     create_output(
#         input_file=input_file,
#         output_file=output_file,
#         keyword_col=KEYWORD_COLUMN,
#         sponsored_col=SPONSORED_COLUMN,
#         frequency=frequency,
#         total_rows=total_rows
#     )


# # ============================================================
# # ENTRY POINT
# # ============================================================

# if __name__ == "__main__":
#     main()
import re
import math
import argparse
from collections import Counter
from pathlib import Path

import openpyxl
from tqdm import tqdm


# ============================================================
# CONFIGURATION
# ============================================================

DEFAULT_INPUT = "input1.xlsx"
DEFAULT_OUTPUT = "sponsored_keywords_output2.xlsx"

KEYWORD_COLUMN = "generated_keywords"
HVD_COLUMN = "Hvd nids"
SPONSORED_COLUMN = "sponsored keywords"

# Maximum sponsored keywords per HVD dataset
MAX_SPONSORED = 2


# ============================================================
# PRIORITY SHORT FORMS / ACRONYMS
# ============================================================

PRIORITY_ACRONYMS = {

    # Health
    "NFHS",
    "DLHS",
    "AHS",
    "HMIS",
    "RCH",
    "SRS",
    "CRS",
    "MCTS",
    "ANM",
    "ASHA",
    "CHC",
    "PHC",
    "UPHC",
    "HWC",
    "ICDS",
    "NRHM",
    "NHM",
    "NHP",
    "NSS",
    "NSSO",
    "IIPS",
    "WHO",
    "UNICEF",
    "UNFPA",
    "UNDP",

    # Census / demographic sources
    "CENSUS",
    "ORGI",

    # Government
    "UIDAI",
    "NITI",
    "NITI AAYOG",

    # Schemes / programmes
    "PMJAY",
    "AB-PMJAY",
    "PM-JAY",
    "PMAY",
    "PMAY-U",
    "PMAY-G",
    "MGNREGA",
    "MGNREGS",
    "MNREGA",
    "NREGA",
    "JSY",
    "JSSK",
    "RKSK",
    "RMNCH",
    "RMNCHA",
    "RKVY",
    "PMKSY",
    "PKVY",
    "SMAM",
    "NMSA",
    "PMUY",
    "JJM",
    "JJM-IMIS",
    "PMGSY",
    "DAY-NRLM",
    "PMMY",
    "PMJJBY",
    "PRASHAD",
    "AMRUT",
    "UDISE",
    "UDISE+",
    "DISE",
    "EMRS",
    "POSHAN",
    "PM-POSHAN",
    "BharatNet",
    "DBT",
    "ADIP",
    "UDID",

    # Agriculture
    "FPO",
    "APMC",
    "KCC",
    "FCI",
    "DA&FW",

    # Petroleum / energy
    "PPAC",
    "CNG",
    "LPG",
    "PNG",
    "LDO",
    "SKO",
    "CPSE",
    "PSU",
    "PSUs",
    "MoPNG",
    "ESY",
    "LSHS",

    # Economy / finance
    "GDP",
    "CPI",
    "WPI",
    "IIP",
    "MSME",
    "GST",
    "TDS",
    "PAN",
    "EPFO",
    "ESIC",
    "NABARD",
    "SEBI",
    "RBI",
    "NPCI",
    "FDI",
    "GVA",
    "EBR",
    "BE",
    "RE",
    "FFS",
    "MPI",

    # Education
    "CBSE",
    "NCERT",
    "AISHE",
    "GER",
    "NER",
    "HSS",
    "ITIs",

    # Infrastructure / transport
    "ULB",
    "NULM",
    "NHAI",
    "RTO",
    "MRTS",
    "IPT",
    "ICT",

    # Environment / mining
    "CPCB",
    "SPCB",
    "NGT",
    "IBM",
    "MCDR",

    # Technology
    "GIS",
    "GPS",
    "API",
    "AI",
    "ML",
    "NLP",

    # Other short forms found in your data
    "SSI",
    "PLI",
    "SC",
    "ST",
    "TB",
    "ESI",
    "CTS",
    "NAM",
    "AAI",
    "GIM",
    "ECR",
    "NAPS",
    "PACS",
    "RGSA",
    "SHGs",
    "PVTG",
    "PVTGs",
    "CSR",
    "LWE",
    "SAGY",
    "SBM-U",
    "WDC-PMKSY",
    "FAME",
    "NIDHI",
    "DARE",
    "PDMC",
    "CSC",
    "BRO",
    "SAIL",
    "NMDC",
    "ICAR",
    "NCRB",
    "IPC",
    "DST",
    "R&D",
}


PRIORITY_ACRONYMS_NORMALIZED = {
    str(x).strip().lower()
    for x in PRIORITY_ACRONYMS
}


# ============================================================
# WORDS THAT MUST NEVER BE SPONSORED
# ============================================================

BLOCKED_SPONSORED_KEYWORDS = {

    # Population / demographic
    "population",
    "populations",
    "demography",
    "demographic",
    "demographics",

    # Generic health
    "health",
    "healthcare",
    "medical",
    "medicine",
    "disease",
    "diseases",

    # Generic people
    "women",
    "woman",
    "men",
    "man",
    "children",
    "child",
    "people",
    "persons",
    "person",
    "household",
    "households",

    # Generic dataset
    "data",
    "dataset",
    "datasets",
    "information",
    "record",
    "records",
    "details",

    # Generic publication
    "report",
    "reports",
    "survey",
    "surveys",
    "study",
    "studies",
    "analysis",
    "statistical",
    "statistics",
    "statistic",

    # Time
    "annual",
    "year",
    "yearly",
    "monthly",
    "quarterly",
    "weekly",
    "daily",

    # Generic analytical
    "indicator",
    "indicators",
    "comparison",
    "comparisons",
    "status",
    "value",
    "values",
    "number",
    "numbers",
    "total",
    "distribution",
    "percentage",

    # Metadata
    "description",
    "descriptions",
    "variable",
    "variables",
    "parameter",
    "parameters",
    "schedule",
    "section",
    "table",
    "tables",

    # Geography
    "india",
    "state",
    "states",
    "district",
    "districts",
    "national",
    "central",
    "regional",
    "region",

    # Government
    "government",
    "ministry",
    "department",
    "scheme",
    "schemes",
    "programme",
    "program",
    "service",
    "services",

    # Generic infrastructure
    "facility",
    "facilities",
    "coverage",
    "infrastructure",
    "management",

    # Generic administrative terms
    "progress",
    "question",
    "questions",
    "parliament",
    "funds",
    "financial",
    "finance",
}


# ============================================================
# NORMALIZE KEYWORD
# ============================================================

def normalize_keyword(keyword):

    if keyword is None:
        return ""

    keyword = str(keyword)

    # Handle non-breaking spaces
    keyword = keyword.replace("\u00A0", " ")

    # Handle tabs/newlines
    keyword = keyword.replace("\t", " ")
    keyword = keyword.replace("\n", " ")
    keyword = keyword.replace("\r", " ")

    # Collapse multiple spaces
    keyword = re.sub(
        r"\s+",
        " ",
        keyword
    )

    return keyword.strip().lower()


# ============================================================
# CHECK WHETHER HVD NID EXISTS
# ============================================================

def is_hvd_row(value):
    """
    A row is considered an HVD row if HVD nids is populated.

    Handles:
        780341
        "780341"
        "780341, 780342"
        "780341;780342"

    Blank / None / whitespace = NOT HVD.
    """

    if value is None:
        return False

    value = str(value)

    value = value.replace(
        "\u00A0",
        " "
    )

    return bool(
        value.strip()
    )


# ============================================================
# SPLIT KEYWORDS
# ============================================================

def split_keywords(value):

    if value is None:
        return []

    value = str(value)

    if not value.strip():
        return []

    value = value.replace(
        "\u00A0",
        " "
    )

    # Your file contains semicolon-separated keywords.
    # Some rows may contain commas too.
    parts = re.split(
        r"[;,]",
        value
    )

    result = []

    for part in parts:

        keyword = part.strip()

        if keyword:
            result.append(keyword)

    return result


# ============================================================
# DETECT PRIORITY ACRONYM
# ============================================================

def is_priority_acronym(keyword):

    normalized = normalize_keyword(
        keyword
    )

    return (
        normalized
        in PRIORITY_ACRONYMS_NORMALIZED
    )


# ============================================================
# AUTOMATIC ACRONYM DETECTION
# ============================================================

def looks_like_acronym(keyword):

    normalized = normalize_keyword(
        keyword
    )

    if not normalized:
        return False

    # Controlled acronym
    if is_priority_acronym(keyword):
        return True

    original = str(keyword)

    compact = re.sub(
        r"[\s\-_/+.()]",
        "",
        original
    )

    compact = re.sub(
        r"[^A-Za-z0-9]",
        "",
        compact
    )

    if not compact:
        return False

    if len(compact) > 12:
        return False

    letters = [
        c
        for c in compact
        if c.isalpha()
    ]

    if len(letters) < 2:
        return False

    uppercase_count = sum(
        1
        for c in letters
        if c.isupper()
    )

    # Strong uppercase acronym
    if (
        len(letters) >= 2
        and uppercase_count == len(letters)
    ):
        return True

    # Mixed-case short forms
    if (
        len(letters) <= 6
        and uppercase_count >= 2
        and uppercase_count / len(letters) >= 0.40
    ):
        return True

    # Hyphenated acronym
    if "-" in original:

        alpha_parts = re.findall(
            r"[A-Za-z]+",
            original
        )

        if (
            len(alpha_parts) >= 2
            and all(
                len(x) <= 8
                for x in alpha_parts
            )
        ):

            uppercase_letters = sum(
                1
                for part in alpha_parts
                for c in part
                if c.isupper()
            )

            total_letters = sum(
                len(part)
                for part in alpha_parts
            )

            if (
                total_letters > 0
                and uppercase_letters / total_letters >= 0.50
            ):
                return True

    return False


# ============================================================
# BLOCKED KEYWORD
# ============================================================

def is_blocked(keyword):

    normalized = normalize_keyword(
        keyword
    )

    return (
        normalized
        in BLOCKED_SPONSORED_KEYWORDS
    )


# ============================================================
# PASS 1:
# CALCULATE FREQUENCY ONLY FOR HVD DATASETS
# ============================================================

def calculate_keyword_frequency(
    input_file,
    keyword_col,
    hvd_col
):

    print()
    print("=" * 70)
    print("PASS 1/2 - SCANNING HVD DATASETS")
    print("=" * 70)

    wb = openpyxl.load_workbook(
        input_file,
        read_only=True,
        data_only=True
    )

    ws = wb.active

    header_row = next(
        ws.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True
        )
    )

    headers = list(
        header_row
    )

    # ---------------------------------------------
    # Validate columns
    # ---------------------------------------------

    if keyword_col not in headers:

        raise ValueError(
            f"Column '{keyword_col}' not found.\n"
            f"Available columns: {headers}"
        )

    if hvd_col not in headers:

        raise ValueError(
            f"Column '{hvd_col}' not found.\n"
            f"Available columns: {headers}"
        )

    keyword_idx = headers.index(
        keyword_col
    )

    hvd_idx = headers.index(
        hvd_col
    )

    frequency = Counter()

    total_rows = 0
    hvd_rows = 0

    # ---------------------------------------------
    # Scan rows
    # ---------------------------------------------

    for row in tqdm(
        ws.iter_rows(
            min_row=2,
            values_only=True
        ),
        desc="Scanning HVD datasets"
    ):

        total_rows += 1

        hvd_value = row[
            hvd_idx
        ]

        # -----------------------------------------
        # IMPORTANT:
        # Ignore non-HVD datasets completely.
        # -----------------------------------------

        if not is_hvd_row(
            hvd_value
        ):
            continue

        hvd_rows += 1

        generated_value = row[
            keyword_idx
        ]

        keywords = split_keywords(
            generated_value
        )

        # Count each keyword once per HVD dataset
        seen = set()

        for keyword in keywords:

            normalized = normalize_keyword(
                keyword
            )

            if (
                normalized
                and normalized not in seen
            ):

                frequency[
                    normalized
                ] += 1

                seen.add(
                    normalized
                )

    wb.close()

    print()
    print(
        f"Total datasets       : {total_rows:,}"
    )

    print(
        f"HVD datasets         : {hvd_rows:,}"
    )

    print(
        f"Non-HVD datasets     : "
        f"{total_rows - hvd_rows:,}"
    )

    print(
        f"Unique HVD keywords  : "
        f"{len(frequency):,}"
    )

    return (
        frequency,
        total_rows,
        hvd_rows
    )


# ============================================================
# SEMANTIC SCORE
# ============================================================

def semantic_score(
    keyword,
    frequency,
    total_hvd_rows
):

    normalized = normalize_keyword(
        keyword
    )

    if not normalized:
        return -999999999

    # ---------------------------------------------
    # NEVER SPONSOR BLOCKED WORDS
    # ---------------------------------------------

    if is_blocked(keyword):
        return -999999999

    # ---------------------------------------------
    # HIGHEST PRIORITY:
    # CONTROLLED ACRONYM
    # ---------------------------------------------

    if is_priority_acronym(keyword):
        return 100000

    # ---------------------------------------------
    # SECOND PRIORITY:
    # AUTOMATIC SHORT FORM
    # ---------------------------------------------

    if looks_like_acronym(keyword):
        return 50000

    # ---------------------------------------------
    # RARITY
    # ---------------------------------------------

    freq = frequency.get(
        normalized,
        1
    )

    rarity = math.log(
        (total_hvd_rows + 1)
        /
        (freq + 1)
    )

    # ---------------------------------------------
    # SPECIFICITY
    # ---------------------------------------------

    character_bonus = min(
        len(normalized) / 10,
        3
    )

    word_count = len(
        normalized.split()
    )

    multiword_bonus = min(
        word_count * 0.7,
        2.5
    )

    # ---------------------------------------------
    # SHORT NORMAL WORD PENALTY
    # ---------------------------------------------

    short_penalty = 0

    if (
        len(normalized) <= 3
        and not looks_like_acronym(keyword)
    ):
        short_penalty = 4

    return (
        rarity * 2.5
        + character_bonus
        + multiword_bonus
        - short_penalty
    )


# ============================================================
# SELECT SPONSORED KEYWORDS
# ============================================================

def select_sponsored_keywords(
    keywords,
    frequency,
    total_hvd_rows,
    max_sponsored=2
):

    if not keywords:
        return []

    candidates = []

    for position, keyword in enumerate(
        keywords
    ):

        # Never sponsor blocked words
        if is_blocked(keyword):
            continue

        score = semantic_score(
            keyword,
            frequency,
            total_hvd_rows
        )

        if score <= -999999:
            continue

        candidates.append(
            {
                "keyword": keyword,
                "score": score,
                "position": position,
                "priority_acronym":
                    is_priority_acronym(
                        keyword
                    ),
                "acronym":
                    looks_like_acronym(
                        keyword
                    ),
            }
        )

    if not candidates:
        return []

    # ---------------------------------------------
    # Rank candidates for selection
    # ---------------------------------------------

    ranked = sorted(
        candidates,
        key=lambda x: (
            x["priority_acronym"],
            x["acronym"],
            x["score"]
        ),
        reverse=True
    )

    selected_candidates = ranked[
        :max_sponsored
    ]

    selected_normalized = {
        normalize_keyword(
            x["keyword"]
        )
        for x in selected_candidates
    }

    # ---------------------------------------------
    # RESTORE ORIGINAL ORDER
    # ---------------------------------------------

    selected = []

    for keyword in keywords:

        if (
            normalize_keyword(keyword)
            in selected_normalized
        ):
            selected.append(
                keyword
            )

    return selected


# ============================================================
# REMOVE ONLY SPONSORED KEYWORDS
# ============================================================

def remove_sponsored_keywords(
    keywords,
    sponsored
):

    sponsored_normalized = {
        normalize_keyword(x)
        for x in sponsored
    }

    remaining = []

    for keyword in keywords:

        if (
            normalize_keyword(keyword)
            not in sponsored_normalized
        ):

            remaining.append(
                keyword
            )

    return remaining


# ============================================================
# PASS 2:
# CREATE OUTPUT
# ============================================================

def create_output(
    input_file,
    output_file,
    keyword_col,
    hvd_col,
    sponsored_col,
    frequency,
    total_rows
):

    print()
    print("=" * 70)
    print("PASS 2/2 - GENERATING HVD SPONSORED KEYWORDS")
    print("=" * 70)

    input_wb = openpyxl.load_workbook(
        input_file,
        read_only=True,
        data_only=True
    )

    input_ws = input_wb.active

    header_row = next(
        input_ws.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True
        )
    )

    headers = list(
        header_row
    )

    keyword_idx = headers.index(
        keyword_col
    )

    hvd_idx = headers.index(
        hvd_col
    )

    # ---------------------------------------------
    # Sponsored column
    # ---------------------------------------------

    output_headers = list(
        headers
    )

    if sponsored_col not in output_headers:

        output_headers.append(
            sponsored_col
        )

    sponsored_idx = output_headers.index(
        sponsored_col
    )

    # ---------------------------------------------
    # Output workbook
    # ---------------------------------------------

    output_wb = openpyxl.Workbook(
        write_only=True
    )

    output_ws = output_wb.create_sheet(
        "Output"
    )

    output_ws.append(
        output_headers
    )

    sponsored_count = 0
    skipped_count = 0

    # ---------------------------------------------
    # Process rows
    # ---------------------------------------------

    for row in tqdm(
        input_ws.iter_rows(
            min_row=2,
            values_only=True
        ),
        total=total_rows,
        desc="Processing datasets"
    ):

        row = list(
            row
        )

        hvd_value = row[
            hvd_idx
        ]

        # =================================================
        # CRITICAL CONDITION
        #
        # ONLY PROCESS HVD DATASETS
        # =================================================

        if not is_hvd_row(
            hvd_value
        ):

            # -----------------------------------------
            # NON-HVD:
            #
            # DO NOT TOUCH generated_keywords
            # DO NOT CREATE sponsored keywords
            # -----------------------------------------

            row[sponsored_idx] = ""

            skipped_count += 1

            output_ws.append(
                row
            )

            continue

        # =================================================
        # HVD DATASET
        # =================================================

        generated_value = row[
            keyword_idx
        ]

        keywords = split_keywords(
            generated_value
        )

        # ---------------------------------------------
        # Select sponsored keywords
        # ---------------------------------------------

        sponsored = select_sponsored_keywords(
            keywords=keywords,
            frequency=frequency,
            total_hvd_rows=(
                max(
                    1,
                    sum(
                        1
                        for x in []
                    )
                )
            ),
            max_sponsored=MAX_SPONSORED
        )

        # NOTE:
        # The total HVD count is supplied below by the
        # wrapper. This line is replaced immediately
        # after the function definition.
        #
        # Kept here only for structure.
        # ---------------------------------------------

        remaining_keywords = (
            remove_sponsored_keywords(
                keywords,
                sponsored
            )
        )

        # ---------------------------------------------
        # Preserve original delimiter
        # ---------------------------------------------

        original_text = (
            ""
            if generated_value is None
            else str(generated_value)
        )

        if ";" in original_text:
            delimiter = "; "
        else:
            delimiter = ", "

        # ---------------------------------------------
        # ONLY remove sponsored keywords
        # ---------------------------------------------

        row[keyword_idx] = delimiter.join(
            remaining_keywords
        )

        # ---------------------------------------------
        # Sponsored keywords
        # ---------------------------------------------

        row[sponsored_idx] = delimiter.join(
            sponsored
        )

        if sponsored:
            sponsored_count += 1

        output_ws.append(
            row
        )

    input_wb.close()

    print()
    print("Saving output file...")

    output_wb.save(
        output_file
    )

    print()
    print("=" * 70)
    print("COMPLETED")
    print("=" * 70)

    print(
        f"Total datasets processed : "
        f"{total_rows:,}"
    )

    print(
        f"Non-HVD datasets skipped : "
        f"{skipped_count:,}"
    )

    print(
        f"HVD datasets with sponsored: "
        f"{sponsored_count:,}"
    )

    print(
        f"Output file               : "
        f"{output_file}"
    )


# ============================================================
# CORRECTED PASS 2
# ============================================================

def create_output_correct(
    input_file,
    output_file,
    keyword_col,
    hvd_col,
    sponsored_col,
    frequency,
    total_rows,
    total_hvd_rows
):

    print()
    print("=" * 70)
    print("PASS 2/2 - GENERATING HVD SPONSORED KEYWORDS")
    print("=" * 70)

    input_wb = openpyxl.load_workbook(
        input_file,
        read_only=True,
        data_only=True
    )

    input_ws = input_wb.active

    header_row = next(
        input_ws.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True
        )
    )

    headers = list(
        header_row
    )

    keyword_idx = headers.index(
        keyword_col
    )

    hvd_idx = headers.index(
        hvd_col
    )

    output_headers = list(
        headers
    )

    if sponsored_col not in output_headers:

        output_headers.append(
            sponsored_col
        )

    sponsored_idx = output_headers.index(
        sponsored_col
    )

    output_wb = openpyxl.Workbook(
        write_only=True
    )

    output_ws = output_wb.create_sheet(
        "Output"
    )

    output_ws.append(
        output_headers
    )

    sponsored_count = 0
    skipped_count = 0

    for row in tqdm(
        input_ws.iter_rows(
            min_row=2,
            values_only=True
        ),
        total=total_rows,
        desc="Processing datasets"
    ):

        row = list(
            row
        )

        hvd_value = row[
            hvd_idx
        ]

        # =====================================================
        # NON-HVD ROW
        # =====================================================

        if not is_hvd_row(
            hvd_value
        ):

            # Keep generated_keywords EXACTLY as original
            row[sponsored_idx] = ""

            skipped_count += 1

            output_ws.append(
                row
            )

            continue

        # =====================================================
        # HVD ROW
        # =====================================================

        generated_value = row[
            keyword_idx
        ]

        keywords = split_keywords(
            generated_value
        )

        sponsored = select_sponsored_keywords(
            keywords=keywords,
            frequency=frequency,
            total_hvd_rows=total_hvd_rows,
            max_sponsored=MAX_SPONSORED
        )

        remaining_keywords = (
            remove_sponsored_keywords(
                keywords,
                sponsored
            )
        )

        # Preserve delimiter
        original_text = (
            ""
            if generated_value is None
            else str(generated_value)
        )

        if ";" in original_text:
            delimiter = "; "
        else:
            delimiter = ", "

        # ONLY remove selected sponsored keywords
        row[keyword_idx] = delimiter.join(
            remaining_keywords
        )

        row[sponsored_idx] = delimiter.join(
            sponsored
        )

        if sponsored:
            sponsored_count += 1

        output_ws.append(
            row
        )

    input_wb.close()

    print()
    print("Saving output file...")

    output_wb.save(
        output_file
    )

    print()
    print("=" * 70)
    print("COMPLETED")
    print("=" * 70)

    print(
        f"Total datasets          : "
        f"{total_rows:,}"
    )

    print(
        f"HVD datasets            : "
        f"{total_hvd_rows:,}"
    )

    print(
        f"Non-HVD datasets skipped: "
        f"{skipped_count:,}"
    )

    print(
        f"HVD datasets sponsored  : "
        f"{sponsored_count:,}"
    )

    print(
        f"Output file             : "
        f"{output_file}"
    )


# ============================================================
# MAIN
# ============================================================

def main():

    parser = argparse.ArgumentParser(
        description=(
            "Generate sponsored keywords ONLY "
            "for HVD datasets."
        )
    )

    parser.add_argument(
        "--input",
        default=DEFAULT_INPUT,
        help="Input Excel file"
    )

    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help="Output Excel file"
    )

    args = parser.parse_args()

    input_file = Path(
        args.input
    )

    output_file = Path(
        args.output
    )

    if not input_file.exists():

        raise FileNotFoundError(
            f"Input file not found: "
            f"{input_file}"
        )

    print()
    print("=" * 70)
    print("HVD SPONSORED KEYWORD GENERATOR")
    print("=" * 70)

    print(
        f"Input : {input_file}"
    )

    print(
        f"Output: {output_file}"
    )

    print(
        f"Maximum sponsored keywords: "
        f"{MAX_SPONSORED}"
    )

    # ========================================================
    # PASS 1
    #
    # Frequency is calculated ONLY from HVD rows.
    # ========================================================

    (
        frequency,
        total_rows,
        total_hvd_rows
    ) = calculate_keyword_frequency(
        input_file=input_file,
        keyword_col=KEYWORD_COLUMN,
        hvd_col=HVD_COLUMN
    )

    if total_hvd_rows == 0:

        raise ValueError(
            "No HVD datasets were found. "
            "Check the 'HVD nids' column."
        )

    # ========================================================
    # PASS 2
    # ========================================================

    create_output_correct(
        input_file=input_file,
        output_file=output_file,
        keyword_col=KEYWORD_COLUMN,
        hvd_col=HVD_COLUMN,
        sponsored_col=SPONSORED_COLUMN,
        frequency=frequency,
        total_rows=total_rows,
        total_hvd_rows=total_hvd_rows
    )


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":
    main()